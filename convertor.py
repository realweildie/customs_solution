import openpyxl

# this file lets you convert 1C store file to necessary format (format of packing list) 

wb = openpyxl.load_workbook("store.xlsx")
sheet = wb.active
max_row = sheet.max_row

for i in range(1, max_row):
    vendor = sheet.cell(row=i, column=1).value

    if any(map(str.isdigit, str(vendor))):
        chain_links = vendor.split()

        vendor_chain = (
            chain_links[0] + "-" + chain_links[-2][:-1] + " р." + chain_links[-3]
        )

        print(vendor_chain)
