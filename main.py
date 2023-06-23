import os, datetime
import openpyxl
from openpyxl.styles import PatternFill

# takes all packing lists in one array and just then check all products
# 60 products and 5 packing lists (320 products) takes 1.2s
# next test on 1900 products and 121 packing lists (quantity is unknown yet)

# paths
PATH_TO_MANY = "docs/"
PATH_TO_COMAPRE_FILES = "output/"

# WARNING: All row constants are actual just for Quiksilver packing lists
# ROW constatns
RETURN_ROW_START = 5
COMPARE_ROW_START = 19
END_USELESS_ROWS_TO_COMPARE = 15
INVOICE_NUMBER = "G14"
INVOICE_DATE = "I14"

# colors
RED = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
GREEN = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

# files for comparing
directory = list(os.listdir(PATH_TO_MANY))
files_for_comparing = list(filter(lambda file: file[-5:] == ".xlsx", directory))

# vendoe cnain | quantity | coordinate
result = [[] for _ in range(len(files_for_comparing))]
invoices = []

start_time = datetime.datetime.now()

# conver files to big array
for i in range(len(files_for_comparing)):
    workbook_compare_obj = openpyxl.load_workbook(PATH_TO_MANY + files_for_comparing[i])
    sheet_compare_obj = workbook_compare_obj.active
    max_comapre_row = sheet_compare_obj.max_row

    invoice_number = sheet_compare_obj[INVOICE_NUMBER].value
    invoice_date = sheet_compare_obj[INVOICE_DATE].value.date()
    invoice_as_string = invoice_number + " от " + str(invoice_date)

    invoices.append(invoice_as_string)

    for j in range(COMPARE_ROW_START, max_comapre_row - END_USELESS_ROWS_TO_COMPARE):
        compare_signature_cell_obj = sheet_compare_obj.cell(row=j, column=2)

        if "DC" in str(compare_signature_cell_obj.value):
            chain_links = compare_signature_cell_obj.value.split(" ")
            vendor_chain = chain_links[2] + " " + chain_links[3]
            quantity = int(float(sheet_compare_obj.cell(row=j, column=11).value))

            result[i].append(
                [vendor_chain, quantity, compare_signature_cell_obj.coordinate]
            )

        if "QUIKSILVER" in str(compare_signature_cell_obj.value) or "BILLABONG" in str(
            compare_signature_cell_obj.value
        ):
            chain_links = compare_signature_cell_obj.value.split(" ")
            vendor_chain = chain_links[1] + " " + chain_links[2]
            quantity = int(float(sheet_compare_obj.cell(row=j, column=11).value))

            result[i].append(
                [vendor_chain, quantity, compare_signature_cell_obj.coordinate]
            )


found_cells = [[] for _ in range(len(files_for_comparing))]

workbook_return_obj = openpyxl.load_workbook("return.xlsx")
sheet_return_obj = workbook_return_obj.active
max_return_row = sheet_return_obj.max_row

# open return file
for i in range(RETURN_ROW_START, max_return_row + 1):
    vendorcode_cell_obj = sheet_return_obj.cell(row=i, column=3)
    color_cell_obj = sheet_return_obj.cell(row=i, column=7)
    size_cell_obj = sheet_return_obj.cell(row=i, column=5)
    quantity_cell_obj = sheet_return_obj.cell(row=i, column=9)
    invoice_cell_obj = sheet_return_obj.cell(row=i, column=11)

    is_invoice_cell_changed = False
    found_quantity_cell = sheet_return_obj.cell(row=i, column=1)

    full_vendorcode = (
        vendorcode_cell_obj.value
        + "-"
        + color_cell_obj.value
        + " р."
        + size_cell_obj.value
    )

    total_found = 0
    for i in range(len(result)):
        for j in range(len(result[i])):
            if full_vendorcode == result[i][j][0]:
                found_cells[i].append(result[i][j][2])
                total_found += result[i][j][1]  # quantity
                found_quantity_cell.value = total_found

                if is_invoice_cell_changed:
                    invoice_cell_obj.value += ", " + invoices[i]
                else:
                    invoice_cell_obj.value = invoices[i]
                    is_invoice_cell_changed = True

                # in return file
                if found_quantity_cell.value >= quantity_cell_obj.value:
                    vendorcode_cell_obj.fill = GREEN
                else:
                    vendorcode_cell_obj.fill = YELLOW

    if total_found == 0:
        vendorcode_cell_obj.fill = RED

    total_found = 0

workbook_return_obj.save("result.xlsx")

# apply changes
for i in range(len(files_for_comparing)):
    workbook_apply_obj = openpyxl.load_workbook(PATH_TO_MANY + files_for_comparing[i])
    sheet_apply_obj = workbook_apply_obj.active

    for j in range(len(found_cells[i])):
        sheet_apply_obj[found_cells[i][j]].fill = GREEN

    new_compare_object_title = files_for_comparing[i][:-5] + "_new.xlsx"
    workbook_apply_obj.save(PATH_TO_COMAPRE_FILES + new_compare_object_title)
    workbook_apply_obj.close()


print("Time elapsed: ", datetime.datetime.now() - start_time)
